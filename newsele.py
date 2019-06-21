# This code is an example of web crawling using selenium and storing that data into a excel file by using openpyxl module
#do the necessary chances according to your need 

from selenium import webdriver
from openpyxl import workbook
import openpyxl
path=r"C:\Users\SuperAdmin\Desktop\test1.xlsx"#path of the excel file 
drive = webdriver.Chrome()
drive.get("https://www.vouchercodes.co.uk/blog/")#just an example paste the url pf any website 


aut = drive.find_elements_by_class_name('byline')# insert class name
cat = drive.find_elements_by_class_name('category')
ti = drive.find_elements_by_class_name('tp-medium a')

work = openpyxl.load_workbook(path)
she = work.active


a = len(aut)
print(a)
b = []
for elements in aut:
	abc = elements.text
	b.append(abc)
print(b)
for i in b:
	for r in range(1,len(b)):
		she.cell(row= r+1, column = 1).value= b[r] 


c = len(cat)
print(c)
d=[]
for element1 in cat:
	cde = element1.text
	d.append(cde)
print(d)

for j in b:
	for k in range(0, len(d)):
		she.cell(row = k+2, column = 2).value = d[k]	

e = len(ti)
print(e)
f=[]
for element2 in ti:
	fgh = element2.text
	f.append(fgh)
print(f)

for q in f:
	for z in range(0, len(f)):
		she.cell(row = z+2, column = 3).value = f[z]	



work.save(path)	
	
drive.close()

